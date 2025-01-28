import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill
import io

def generate_add_edit_employee_using_excel_template(user):
    # Define the fields for the template
    data = []

    columns = [
        "Paycode (required)",
        "Attendance Card Number (required)",
        "Employee Name (required)",
        "Father's/Husband's Name",
        "Mother's Name",
        "Wife's Name",
        "Date of Birth",
        "Phone Number",
        "Alternate Phone Number",
        "Pan Number",
        "Local District",
        "Religion",
        "Driving Licence",
        "Local State or UT (dropdown)",
        "Email",
        "Passport",
        "Local Pin Code",
        "Aadhaar",
        "Handicapped (checkbox)",
        "Gender (dropdown)",
        "Marital Status (dropdown)",
        "Blood Group (dropdown)",
        "Nationality (default: Indian)",
        "Educational Qualification (dropdown)",
        "Technical Qualification",
        "Permanent and Local Same? (checkbox)",
        "Permanent Address",
        "Permanent District",
        "Permanent State or UT (dropdown)",
        "Permanent Pin Code",
        "Local Address",
    ]

    ##Adding Type of Employee Detail Heading (like Personal, Professional, Salary etc)
    personal_detail_heading_dict = ["Employee", "Personal", "Detials"] + ['' for _ in columns[3:]]

    # Define the "Employee Personal Detail" row
    #employee_personal_detail = ["Employee", "Personal", "Detail"] + ['' for _ in range(len(columns) - 3)]
    
    # Create a DataFrame with "Employee Personal Detail" as the first row
    df = pd.DataFrame([personal_detail_heading_dict, columns], columns=columns)
    #df = pd.DataFrame(columns, columns=employee_personal_detail)

   
    # Create the HTTP response for Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="employee_template.xlsx"'
    
    excel_buffer = io.BytesIO()
    # Write the DataFrame to an Excel filegg
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')

        # Get the worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Define the light brown fill
        light_blue_fill = PatternFill(start_color="92c7d4", end_color="92c7d4", fill_type="solid")  # Light blue
        light_red_fill = PatternFill(start_color="f5bac1", end_color="f5bac1", fill_type="solid")  # Light red 

        # Adjust the width of all columns
        # for column in worksheet.columns:
        #     max_length = max(len(str(cell.value)) for cell in column) + 2  # Adding a little extra padding
        #     worksheet.column_dimensions[column[0].column_letter].width = max_length
        #
        # # Get the last row number
        # last_row = len(df) + 1

        # Define the light green fill
        light_green_fill= PatternFill(start_color="b8e3c5", end_color="b8e3c5", fill_type="solid")

        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column) + 2  # Adding a little extra padding
            worksheet.column_dimensions[column[0].column_letter].width = max_length

        # Apply bold font and blue fill color to the last row
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)  # Headers are in the first row
            cell.fill=light_green_fill 
            cell = worksheet.cell(row=2, column=col)
            cell.fill=light_green_fill
            # cell = worksheet.cell(row=last_row, column=col)
            # cell.fill = light_red_fill
            # cell.font = Font(bold=True, color='080659')  # Set the font color to blue

    # Create a response with the Excel file content
    response = HttpResponse(content=excel_buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pf_statement.xlsx"'

    # Close the buffer
    excel_buffer.close()


    return response
