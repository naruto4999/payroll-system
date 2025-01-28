import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from api.models import EmployeeSalaryPrepared

def generate_payment_sheet_as_per_compliance_xlsx(user, request_data, employee_salaries):
    # Initialize a workbook and a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Payment Sheet as per Compliance"

    # Define the headers
    headers = [
        "S/N", "ACN", "Employee Name", "F/H Name", "Designation", "Salary Rate", 
        "Paid Days", "Earned Salary", "Incen.", "OT Hrs", "OT Amt", "Total Earned", 
        "Advance", "EPF", "ESI", "LWF", "Others", "TDS", "Total Deductions", 
        "Net Payable", "Bank Pay.", "Other Pay.", "Signat."
    ]
    worksheet.append(headers)

    # Define row colors
    fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Grouping logic
    group_by = request_data.get('group_by', None)
    grand_total = {
        "earned_salary": 0,
        "incentive": 0,
        "ot_amt": 0,
        "total_earnings": 0,
        "total_advance": 0,
        "total_pf": 0,
        "total_esi": 0,
        "total_lwf": 0,
        "total_others": 0,
        "total_tds": 0,
        "total_deductions": 0,
        "net_payable": 0,
        "bank_payment": 0,
        "other_payment": 0
    }

    # Process employee salaries
    for index, salary in enumerate(employee_salaries):
        row = [
            index + 1,
            salary.employee.attendance_card_no,
            salary.employee.name,
            salary.father_or_husband_name or "",
            salary.employee.professional_detail.designation.name if salary.employee.professional_detail else "",
            salary.salary_rate,
            salary.paid_days,
            0,  # Placeholder for earned_salary
            0,  # Placeholder for incentive
            0,  # Placeholder for ot_hours
            0,  # Placeholder for ot_amount
            0,  # Placeholder for total_earnings
            0,  # Placeholder for advance
            0,  # Placeholder for epf
            0,  # Placeholder for esi
            0,  # Placeholder for lwf
            0,  # Placeholder for others
            0,  # Placeholder for tds
            0,  # Placeholder for total_deductions
            0,  # Placeholder for net_payable
            0,  # Placeholder for bank_payment
            0,  # Placeholder for other_payment
            ""
        ]

        # Calculate values based on existing logic
        total_earnings_amount = sum(earned.earned_amount for earned in salary.current_salary_earned_amounts.all())
        row[7] = total_earnings_amount  # Earned Salary
        grand_total["earned_salary"] += total_earnings_amount

        row[8] = salary.incentive_amount  # Incentive
        grand_total["incentive"] += salary.incentive_amount

        row[9] = salary.net_ot_minutes_monthly / 60 if salary.net_ot_minutes_monthly else 0  # OT Hrs
        row[10] = salary.net_ot_amount_monthly  # OT Amount
        grand_total["ot_amt"] += row[10]

        row[11] = total_earnings_amount + row[10] + row[8]  # Total Earnings
        grand_total["total_earnings"] += row[11]

        row[12] = salary.advance  # Advance
        grand_total["total_advance"] += row[12]

        row[13] = salary.epf  # EPF
        grand_total["total_pf"] += row[13]

        row[14] = salary.esi  # ESI
        grand_total["total_esi"] += row[14]

        row[15] = salary.lwf  # LWF
        grand_total["total_lwf"] += row[15]

        row[16] = salary.others  # Others
        grand_total["total_others"] += row[16]

        row[17] = salary.tds  # TDS
        grand_total["total_tds"] += row[17]

        # Total Deductions
        total_deductions = salary.advance + salary.epf + salary.esi + salary.tds + salary.others
        row[18] = total_deductions
        grand_total["total_deductions"] += total_deductions

        # Net Payable
        row[19] = row[11] - total_deductions
        grand_total["net_payable"] += row[19]

        # Bank Payment
        row[20] = row[19] - row[12]  # Assuming other_payment is calculated as net_payable - bank_payment
        grand_total["bank_payment"] += row[20]

        worksheet.append(row)

    # Append grand total row
    total_row = ["Grand Total", "", "", "", "", "", "", 
                 grand_total["earned_salary"], grand_total["incentive"], 
                 grand_total["ot_amt"], grand_total["total_earnings"], 
                 grand_total["total_advance"], grand_total["total_pf"], 
                 grand_total["total_esi"], grand_total["total_lwf"], 
                 grand_total["total_others"], grand_total["total_tds"], 
                 grand_total["total_deductions"], grand_total["net_payable"], 
                 grand_total["bank_payment"], ""]
    worksheet.append(total_row)

    # Apply fill color to the total row
    for cell in worksheet[-1]:
        cell.fill = fill_color

    # Save the workbook
    file_path = f"payment_sheet_as_per_compliance_{user.id}.xlsx"
    workbook.save(file_path)
    return file_path

