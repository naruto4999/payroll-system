import pandas as pd
from django.http import HttpResponse
import io
from datetime import date, timedelta, datetime
from dateutil.relativedelta import relativedelta
from decimal import Decimal, ROUND_HALF_UP, ROUND_CEILING
import calendar
from api.models import EarningsHead, EmployeeSalaryEarning, PfEsiSetup
from openpyxl.styles import Font, PatternFill


# def calculate_age(date_of_birth, reference_date):
#     # Calculate age based on the provided date_of_birth and reference_date
#     age = reference_date.year - date_of_birth.year - ((reference_date.month, reference_date.day) < (date_of_birth.month, date_of_birth.day))
#     return age


def generate_payment_sheet_xlsx(user, request_data, employee_salaries):
    company_pf_esi_setup = PfEsiSetup.objects.get(company=request_data['company'])
    employees_data = []
    total_row = {
    'S/N': 'Grand Total',
    'ACN': '',
    'Employee Name': '',
    'Father/Husband Name': '',
    'Designation': '',
    'Salary Rate': 0,
    'PD': 0,
    'Earned Salary': 0,
    'Incentive': 0,
    'OT Hrs': 0,
    'OT Amount': 0,
    'Total Earned': 0,
    'Advance': 0,
    'EPF': 0,
    'ESI': 0,
    'LWF': 0,
    'Others': 0,
    'TDS': 0,
    'Total Deduction': 0,
    'Net Payable': 0
    }
    dept_total_dict = {
        "total_paid_days": 0,
        "total_salary_rate": 0,
        "ot_hrs": 0,
        "earned_salary" : 0,
        "ot_amt": 0,
        "total_earnings": 0,
        'total_incentive': 0,
        "total_advance": 0,
        "total_pf": 0,
        "total_esi": 0,
        "total_lwf": 0,
        "total_others": 0,
        "total_tds": 0,
        "total_deductions": 0,
        "net_payable": 0,
    }
         
    for salary_index, employee_salary in enumerate(employee_salaries):
        if request_data['filters']['group_by'] != 'none':
            # try:
            print(f'printing department name fo index: {salary_index}')
            current_employee_department = employee_salary.employee.employee_professional_detail.department
            previous_employee_department = employee_salaries[salary_index-1].employee.employee_professional_detail.department if salary_index!=0 else None

            
            if salary_index == 0 or current_employee_department != previous_employee_department:
            #if salary_index == 0 or employee_salary.employee.employee_professional_detail.department.name != employee_salaries[salary_index-1].employee.employee_professional_detail.department.name:
                current_employee_department_name = current_employee_department.name if current_employee_department != None else 'No Department'
                departement_name_dict = {
                'S/N': f"{current_employee_department_name} dept_name_finder",
                'ACN': '',
                'Employee Name': '',
                'Father/Husband Name': '',
                'Designation': '',
                'Salary Rate': '',
                'PD': '',
                'Earned Salary': '',
                'Incentive': '',
                'OT Hrs': '',
                'OT Amount': '',
                'Total Earned': '',
                'Advance': '',
                'EPF': '',
                'ESI': '',
                'LWF': '',
                'Others': '',
                'TDS': '',
                'Total Deduction': '',
                'Net Payable': ''
                }

                employees_data.append(departement_name_dict)


                dept_total_dict = {
                    "total_paid_days": 0,
                    "total_salary_rate": 0,
                    "ot_hrs": 0,
                    "earned_salary" : 0,
                    "incentive": 0,
                    "ot_amt": 0,
                    "total_earnings": 0,
                    "total_advance": 0,
                    "total_pf": 0,
                    "total_esi": 0,
                    "total_lwf": 0,
                    "total_others": 0,
                    "total_tds": 0,
                    "total_deductions": 0,
                    "net_payable": 0,
                }
            # except:
            #     print('in the except block ')
            #     pass
        current_employee_dict = {
            'S/N': salary_index+1,
            'ACN': employee_salary.employee.attendance_card_no,
            'Employee Name': employee_salary.employee.name,
            'Father/Husband Name': employee_salary.employee.father_or_husband_name,
            'Designation': '',
            'Salary Rate': 0,
            'PD': 0,
            'Earned Salary': 0,
            'Incentive': 0,
            'OT Hrs': 0,
            'OT Amount': 0,
            'Total Earned': 0,
            'Advance': 0,
            'EPF': 0,
            'ESI': 0,
            'LWF': 0,
            'Others': 0,
            'TDS': 0,
            'Total Deduction': 0,
            'Net Payable': 0
            }
        #Designation
        try: current_employee_dict['Designation'] = employee_salary.employee.employee_professional_detail.designation.name
        except: pass

        #Salary Rate
        try:
            total_earnings_rate = 0
            earnings_heads = EarningsHead.objects.filter(company=employee_salary.company, user=employee_salary.user if employee_salary.user.role == "OWNER" else employee_salary.user.regular_to_owner.owner)
            employee_salary_rates = EmployeeSalaryEarning.objects.filter(employee=employee_salary.employee, from_date__lte=employee_salary.date, to_date__gte=employee_salary.date)
            for head in earnings_heads:
                salary_for_particular_earning_head = employee_salary_rates.filter(earnings_head=head)
                if salary_for_particular_earning_head.exists():
                    total_earnings_rate += salary_for_particular_earning_head.first().value
            current_employee_dict['Salary Rate'] = total_earnings_rate
            total_row['Salary Rate'] += total_earnings_rate
            dept_total_dict['total_salary_rate'] += total_earnings_rate
        except: 
            pass

        #Paid Days
        employee_monthly_attendance_details = None
        try: 
            employee_monthly_attendance_details = employee_salary.employee.monthly_attendance_details.filter(date=employee_salary.date).first()
            current_employee_dict['PD']= employee_monthly_attendance_details.paid_days_count/2
            total_row['PD'] += employee_monthly_attendance_details.paid_days_count/2
            dept_total_dict['total_paid_days'] += employee_monthly_attendance_details.paid_days_count/2

        except: pass        

        #Earned Salary
        try:
            earned_amounts = employee_salary.current_salary_earned_amounts.all()
            #Total Earned
            if earned_amounts and earned_amounts.exists():
                total_earnings_amount = 0
                for earned in earned_amounts:
                    total_earnings_amount += (earned.earned_amount)
            #total_earnings_amount += employee_salary.incentive_amount
            current_employee_dict['Earned Salary'] = total_earnings_amount
            total_row['Earned Salary'] += total_earnings_amount
            dept_total_dict['earned_salary'] += total_earnings_amount
        except: pass

        #Incentive
        try:
            total_earnings_amount += employee_salary.incentive_amount
            current_employee_dict['Incentive'] = employee_salary.incentive_amount
            total_row['Incentive'] += employee_salary.incentive_amount
            dept_total_dict['incentive'] += employee_salary.incentive_amount
        except Exception as e:
            print(f"Error occurred in generate_payment_sheet_xlsx in Incentive: {str(e)}")

        #OT
        ot_hrs = employee_salary.net_ot_minutes_monthly/60
        current_employee_dict['OT Hrs'] = ot_hrs
        total_row['OT Hrs'] += ot_hrs
        dept_total_dict['ot_hrs'] += ot_hrs
        
        #OT Amt
        ot_amt = employee_salary.net_ot_amount_monthly
        current_employee_dict['OT Amount'] = ot_amt
        total_row['OT Amount'] += ot_amt
        dept_total_dict['ot_amt'] += ot_amt

        #Total Earned
        print(f"{employee_salary.employee.name}: {total_earnings_amount}")
        
        total_earned = total_earnings_amount
        if ot_amt:
            total_earned += ot_amt
        current_employee_dict['Total Earned'] = total_earned
        total_row['Total Earned'] += total_earned
        dept_total_dict['total_earnings'] += total_earned

        #Advance
        try:
            advance = employee_salary.advance_deducted
            current_employee_dict['Advance'] = advance
            total_row['Advance'] += advance
            dept_total_dict['total_advance'] += advance
        except: pass

        #EPF
        try:
            pf_vpf_deducted = 0
            pf_vpf_deducted += employee_salary.pf_deducted if employee_salary.pf_deducted !=None else 0
            pf_vpf_deducted += employee_salary.vpf_deducted if employee_salary.vpf_deducted !=None else 0
            current_employee_dict['EPF'] = pf_vpf_deducted
            total_row['EPF'] += pf_vpf_deducted
            dept_total_dict['total_pf'] += pf_vpf_deducted
        except: pass

        #ESI
        try:
            esi = employee_salary.esi_deducted
            current_employee_dict['ESI'] = esi
            total_row['ESI'] += esi
            dept_total_dict['total_esi'] += esi
        except: pass

        #LWF
        if company_pf_esi_setup.enable_labour_welfare_fund:
            if employee_salary.labour_welfare_fund_deducted:
                lwf_deducted = employee_salary.labour_welfare_fund_deducted
                current_employee_dict['LWF'] = lwf_deducted
                total_row['LWF'] += lwf_deducted
                dept_total_dict['total_lwf'] += lwf_deducted

        #TDS
        try:
            tds = employee_salary.tds_deducted
            current_employee_dict['TDS'] = tds
            total_row['TDS'] += tds
            dept_total_dict['total_tds'] +=tds
        except: pass

        #Others
        try:
            others_deducted = employee_salary.others_deducted
            current_employee_dict['Others'] = others_deducted
            total_row['Others'] += others_deducted
            dept_total_dict['total_others'] += others_deducted
        except: pass


        #Total Deductions
        try:
            total_deductions = advance + pf_vpf_deducted + esi + tds + others_deducted
            if company_pf_esi_setup.enable_labour_welfare_fund:
                if employee_salary.labour_welfare_fund_deducted:
                    total_deductions += lwf_deducted
            current_employee_dict['Total Deduction'] = total_deductions
            total_row['Total Deduction'] += total_deductions
            dept_total_dict['total_deductions'] += total_deductions
        except: pass

        #Net Payable
        try:
            print(f"Net payable {employee_salary.employee.name}: {total_earned} - {total_deductions}")
            net_payable = total_earned - total_deductions
            current_employee_dict['Net Payable'] = net_payable
            total_row['Net Payable'] += net_payable
            dept_total_dict['net_payable'] += net_payable
        except: pass

        employees_data.append(current_employee_dict)

        #Dept Total
        if request_data['filters']['group_by'] != 'none':
            if (salary_index != len(employee_salaries)-1 and employee_salary.employee.employee_professional_detail.department != employee_salaries[salary_index+1].employee.employee_professional_detail.department) or (salary_index == len(employee_salaries)-1 and employee_salary.employee.employee_professional_detail.department):
                department_total_dict = {
                    'S/N': 'Department Total',
                    'ACN': '',
                    'Employee Name': '',
                    'Father/Husband Name': '',
                    'Designation': '',
                    'Salary Rate': dept_total_dict['total_salary_rate'],
                    'PD': dept_total_dict['total_paid_days'],
                    'Earned Salary': dept_total_dict['earned_salary'],
                    'Incentive': dept_total_dict['incentive'],
                    'OT Hrs': dept_total_dict['ot_hrs'],
                    'OT Amount': dept_total_dict['ot_amt'],
                    'Total Earned': dept_total_dict['total_earnings'],
                    'Advance': dept_total_dict['total_advance'],
                    'EPF': dept_total_dict['total_pf'],
                    'ESI': dept_total_dict['total_esi'],
                    'LWF': dept_total_dict['total_lwf'],
                    'Others': dept_total_dict['total_others'],
                    'TDS': dept_total_dict['total_tds'],
                    'Total Deduction': dept_total_dict['total_deductions'],
                    'Net Payable': dept_total_dict['net_payable']
                    }
                employees_data.append(department_total_dict)

                #Printing Empty Line
                empty_line_dict = {
                'S/N': '',
                'ACN': '',
                'Employee Name': '',
                'Father/Husband Name': '',
                'Designation': '',
                'Salary Rate': '',
                'PD': '',
                'Earned Salary': '',
                'Incentive': '',
                'OT Hrs': '',
                'OT Amount': '',
                'Total Earned': '',
                'Advance': '',
                'EPF': '',
                'ESI': '',
                'LWF': '',
                'Others': '',
                'TDS': '',
                'Total Deduction': '',
                'Net Payable': ''
                }
                employees_data.append(empty_line_dict)

                

            
    employees_data.append(total_row)

    df = pd.DataFrame(employees_data)
    if not company_pf_esi_setup.enable_labour_welfare_fund:
        df.drop(columns=['LWF'], inplace=True)
    df.rename(columns={'Earned Salary': 'Earned Salary (*incl of arrear)'}, inplace=True)

    # Create a BytesIO buffer
    excel_buffer = io.BytesIO()

    # Write the DataFrame to the Excel file
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        # Get the worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Define the light brown fill
        light_blue_fill = PatternFill(start_color="92c7d4", end_color="92c7d4", fill_type="solid")  # Light blue
        light_red_fill = PatternFill(start_color="f5bac1", end_color="f5bac1", fill_type="solid")  # Light red 

        # Fill color for department name
        if request_data['filters']['group_by'] != 'none':
            department_total_row_numbers = []
            for row_idx in range(2, len(df) + 2):  # Data rows start from the second row
                first_col_value = worksheet.cell(row=row_idx, column=1).value  # Get the value in the first column
                if first_col_value and isinstance(first_col_value, str):
                    # Split the text by spaces
                    words = first_col_value.split()
                    # Check if the last word is "dept_name_finder"
                    if len(words) > 1:
                        if words[-1] == "dept_name_finder":
                            # Apply light brown fill to the cell
                            worksheet.cell(row=row_idx, column=1).fill = light_blue_fill
                            # Replace the text in the cell (remove "dept_name_finder")
                            new_text = " ".join(words[:-1])  # Remove the last word
                            worksheet.cell(row=row_idx, column=1).value = new_text
                        elif words[-1] == 'Total' and words[0] == 'Department':
                            department_total_row_numbers.append(row_idx)

        # Adjust the width of all columns
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column) + 2  # Adding a little extra padding
            worksheet.column_dimensions[column[0].column_letter].width = max_length

        # Get the last row number
        last_row = len(df) + 1

        # Define the light green fill
        header_fill = PatternFill(start_color="b8e3c5", end_color="b8e3c5", fill_type="solid")

        # Apply bold font and blue fill color to the last row
        for col in range(1, len(df.columns) + 1):
            if request_data['filters']['group_by'] != 'none':
                for dept_total_row_no in department_total_row_numbers:
                    cell = worksheet.cell(row=dept_total_row_no, column=col)
                    cell.fill = light_blue_fill
            cell = worksheet.cell(row=1, column=col)  # Headers are in the first row
            cell.fill = header_fill
            cell = worksheet.cell(row=last_row, column=col)
            cell.fill = light_red_fill
            cell.font = Font(bold=True, color='080659')  # Set the font color to blue

    # Create a response with the Excel file content
    response = HttpResponse(content=excel_buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pf_statement.xlsx"'

    # Close the buffer
    excel_buffer.close()

    # Return the response
    return response
